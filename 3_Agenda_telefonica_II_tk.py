import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook

class AgendaTelefonica:
    def __init__(self):
        self.contatos = {}

    def adicionar_contato(self, nome, telefone, tipo_telefone, categoria):
        if nome in self.contatos:
            return f"O contato '{nome}' já existe."
        else:
            self.contatos[nome] = {
                'telefone': telefone,
                'tipo_telefone': tipo_telefone,
                'categoria': categoria
            }
            return f"Contato '{nome}' adicionado com sucesso."

    def remover_contato(self, nome):
        if nome in self.contatos:
            del self.contatos[nome]
            return f"Contato '{nome}' removido com sucesso."
        else:
            return f"Contato '{nome}' não encontrado."

    def pesquisar_contato(self, nome):
        if nome in self.contatos:
            contato = self.contatos[nome]
            return f"Nome: {nome}\nTelefone: {contato['telefone']} ({contato['tipo_telefone']})\nCategoria: {contato['categoria']}"
        else:
            return f"Contato '{nome}' não encontrado."

    def listar_contatos(self):
        if not self.contatos:
            return "A agenda está vazia."
        else:
            contatos_listados = ""
            for nome, dados in self.contatos.items():
                contatos_listados += f"Nome: {nome}\nTelefone: {dados['telefone']} ({dados['tipo_telefone']})\nCategoria: {dados['categoria']}\n\n"
            return contatos_listados

class InterfaceAgendaTelefonica:
    def __init__(self, root):
        self.agenda = AgendaTelefonica()
        self.root = root
        self.root.title("Agenda Telefônica")
        self.create_widgets()

    def create_widgets(self):
        # Adicionar Contato
        self.label_nome = tk.Label(self.root, text="Nome:")
        self.label_nome.grid(row=0, column=0, padx=10, pady=10)
        self.entry_nome = tk.Entry(self.root)
        self.entry_nome.grid(row=0, column=1, padx=10, pady=10)

        self.label_telefone = tk.Label(self.root, text="Telefone: (Formato: DDD (XX) XXXX-XXXX)")
        self.label_telefone.grid(row=1, column=0, padx=10, pady=10)
        self.entry_telefone = tk.Entry(self.root)
        self.entry_telefone.grid(row=1, column=1, padx=10, pady=10)

        # Radio Buttons para Tipo de Telefone
        self.tipo_telefone_var = tk.StringVar()
        self.tipo_telefone_var.set("Celular")  # Valor padrão inicial é Celular

        self.label_tipo_telefone = tk.Label(self.root, text="Tipo de Telefone:")
        self.label_tipo_telefone.grid(row=2, column=0, padx=10, pady=10)

        self.radio_celular = tk.Radiobutton(self.root, text="Celular", variable=self.tipo_telefone_var, value="Celular")
        self.radio_celular.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

        self.radio_fixo = tk.Radiobutton(self.root, text="Fixo", variable=self.tipo_telefone_var, value="Fixo")
        self.radio_fixo.grid(row=2, column=1, padx=10, pady=5, sticky=tk.E)

        # Radio Buttons para Categoria
        self.categoria_var = tk.StringVar()
        self.categoria_var.set("Pessoal")  # Valor padrão inicial é Pessoal

        self.label_categoria = tk.Label(self.root, text="Categoria:")
        self.label_categoria.grid(row=3, column=0, padx=10, pady=10)

        self.radio_pessoal = tk.Radiobutton(self.root, text="Pessoal", variable=self.categoria_var, value="Pessoal")
        self.radio_pessoal.grid(row=3, column=1, padx=10, pady=5, sticky=tk.W)

        self.radio_trabalho = tk.Radiobutton(self.root, text="Trabalho", variable=self.categoria_var, value="Trabalho")
        self.radio_trabalho.grid(row=3, column=1, padx=10, pady=5, sticky=tk.E)

        self.button_adicionar = tk.Button(self.root, text="Adicionar Contato", command=self.adicionar_contato)
        self.button_adicionar.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

        # Remover Contato
        self.label_nome_remover = tk.Label(self.root, text="Nome para Remover:")
        self.label_nome_remover.grid(row=5, column=0, padx=10, pady=10)
        self.entry_nome_remover = tk.Entry(self.root)
        self.entry_nome_remover.grid(row=5, column=1, padx=10, pady=10)

        self.button_remover = tk.Button(self.root, text="Remover Contato", command=self.remover_contato)
        self.button_remover.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

        # Pesquisar Contato
        self.label_nome_pesquisar = tk.Label(self.root, text="Nome para Pesquisar:")
        self.label_nome_pesquisar.grid(row=7, column=0, padx=10, pady=10)
        self.entry_nome_pesquisar = tk.Entry(self.root)
        self.entry_nome_pesquisar.grid(row=7, column=1, padx=10, pady=10)

        self.button_pesquisar = tk.Button(self.root, text="Pesquisar Contato", command=self.pesquisar_contato)
        self.button_pesquisar.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

        # Listar Contatos
        self.button_listar = tk.Button(self.root, text="Listar Contatos", command=self.listar_contatos)
        self.button_listar.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

        # Exportar para Excel
        self.label_diretorio = tk.Label(self.root, text="Diretório para exportar: copie e cole o caminho para salvar o arquivo")
        self.label_diretorio.grid(row=10, column=0, padx=10, pady=10)
        self.entry_diretorio = tk.Entry(self.root, width=50)
        self.entry_diretorio.grid(row=10, column=1, padx=10, pady=10)

        self.button_exportar_excel = tk.Button(self.root, text="Exportar para Excel", command=self.exportar_para_excel)
        self.button_exportar_excel.grid(row=11, column=0, columnspan=2, padx=10, pady=10)

        # Text Area for Displaying Contacts
        self.text_area = tk.Text(self.root, width=50, height=10)
        self.text_area.grid(row=12, column=0, columnspan=2, padx=10, pady=10)

    def adicionar_contato(self):
        nome = self.entry_nome.get()
        telefone = self.entry_telefone.get()
        tipo_telefone = self.tipo_telefone_var.get()
        categoria = self.categoria_var.get()

        mensagem = self.agenda.adicionar_contato(nome, telefone, tipo_telefone, categoria)
        messagebox.showinfo("Informação", mensagem)

    def remover_contato(self):
        nome = self.entry_nome_remover.get()
        mensagem = self.agenda.remover_contato(nome)
        messagebox.showinfo("Informação", mensagem)

    def pesquisar_contato(self):
        nome = self.entry_nome_pesquisar.get()
        mensagem = self.agenda.pesquisar_contato(nome)
        messagebox.showinfo("Informação", mensagem)

    def listar_contatos(self):
        contatos = self.agenda.listar_contatos()
        self.text_area.delete(1.0, tk.END)
        self.text_area.insert(tk.END, contatos)

    def exportar_para_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contatos"

        ws['A1'] = "Nome"
        ws['B1'] = "Telefone"
        ws['C1'] = "Tipo de Telefone"
        ws['D1'] = "Categoria"

        linha = 2
        for nome, dados in self.agenda.contatos.items():
            ws[f'A{linha}'] = nome
            ws[f'B{linha}'] = dados['telefone']
            ws[f'C{linha}'] = dados['tipo_telefone']
            ws[f'D{linha}'] = dados['categoria']
            linha += 1

        # Solicita ao usuário o diretório de destino
        diretorio = self.entry_diretorio.get()
        if not diretorio:
            messagebox.showerror("Erro", "Informe um diretório válido para exportar o arquivo.")
            return

        nome_arquivo = "contatos.xlsx"
        caminho_completo = f"{diretorio}/{nome_arquivo}"

        try:
            wb.save(caminho_completo)
            messagebox.showinfo("Informação", f"Contatos exportados para '{caminho_completo}' com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o arquivo: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = InterfaceAgendaTelefonica(root)
    root.mainloop()
